def BugzillaExcelPop(Reference_file,Updated_File,List_Attributes=[]):
    from Test3func import bugzilla
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(Reference_file)
    updated_wb = openpyxl.load_workbook(Updated_File)
    #print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Sheet1')
    updated_sheet = updated_wb.active
    cell=2
    while (sheet['A'+str(cell)].value!=None):
        updated_sheet.cell(row=cell,column=1).value= sheet['A'+str(cell)].value
        cell=cell+1
    for item in List_Attributes:
        if (item.lower() != "duplicates"):
            cell=2
            while (sheet['A' + str(cell)].value != None):
                updated_sheet.cell(row=cell, column=(List_Attributes.index(item)+2)).value =(bugzilla(sheet['A'+str(cell)].value,str(item.lower())))
                cell = cell + 1
        else:
            cell = 2
            while (sheet['A' + str(cell)].value != None):
                updated_sheet.cell(row=cell, column=(List_Attributes.index(item) + 2)).value =str((bugzilla(sheet['A' + str(cell)].value, str(item.lower()))))
                cell = cell + 1

    updated_wb.save(Updated_File)

    #print("Total ids=",cell-1)
    #print("Total ids(by max_row)=",sheet.max_row)
    #print(sheet['A1'].value)


